import streamlit as st
import imaplib
import email
from email.header import decode_header
import os
import shutil
import json
import re
import glob
import pandas as pd
import numpy as np
import zipfile
import copy
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

# ============================================================
# 1. KONFIGURASI HALAMAN (UI/UX)
# ============================================================
st.set_page_config(
    page_title="Rekap LDK & Transaksi", 
    page_icon="✨", 
    layout="wide",
    initial_sidebar_state="expanded"
)

# File untuk menyimpan memori cabang
CONFIG_FILE = "config_cabang.json"

def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r") as f:
                return json.load(f)
        except:
            pass
    return [{"Nama Cabang": "BANJARMASIN", "Logika Address": "DED-STY-AYN-BJM-CCF"}]

def save_config(data):
    with open(CONFIG_FILE, "w") as f:
        json.dump(data, f)

# ============================================================
# 2. LOGIKA BACKEND & FUNGSI PENDUKUNG
# ============================================================

def clean_filename(filename):
    if not filename: return "untitled_attachment"
    return re.sub(r'[\\/*?:"<>|\n\r]', "", filename).strip()

def decode_mime_header(header_value):
    if not header_value: return ""
    decoded_parts = decode_header(header_value)
    result = ""
    for part, encoding in decoded_parts:
        if isinstance(part, bytes):
            try: result += part.decode(encoding or "utf-8", errors="ignore")
            except: result += part.decode("latin1", errors="ignore")
        else: result += str(part)
    return result

def do_imap_search(mail_conn, query_string):
    words = query_string.split()
    search_criteria = " ".join([f'SUBJECT "{word}"' for word in words])
    status, messages = mail_conn.search(None, search_criteria)
    if status == "OK" and messages[0]: return messages[0].split()
    return []

def extract_date_key(text):
    pattern = r'(?i)(.*?)\s+(\d{1,2})\s+([a-z]+)\s+(\d{4})'
    match = re.search(pattern, text)
    if match: return f"{int(match.group(2))}_{match.group(3).strip().upper()}_{match.group(4).strip()}"
    return None

def merge_downloaded_excel(download_dir):
    excel_files = glob.glob(os.path.join(download_dir, "*.xlsx"))
    valid_files = [f for f in excel_files if "(GABUNGAN)" not in f]
    if len(valid_files) < 2: return None

    data_dfs = []
    header_df = None
    lots_col_idx = 6 
    bs_col_idx = None 
    
    days, months, years, prefixes, summary_rows = [], set(), set(), [], []
    pattern = r'(?i)(.*?)\s+(\d{1,2})\s+([a-z]+)\s+(\d{4})'

    for file_path in valid_files:
        file_name = os.path.basename(file_path)
        match = re.search(pattern, file_name)
        if match:
            prefixes.append(match.group(1).strip())
            day_val = int(match.group(2))
            days.append(day_val)
            month_val = match.group(3).strip().upper()
            months.add(month_val)
            year_val = match.group(4).strip()
            years.add(year_val)
            date_str = f"{day_val:02d}-{month_val[:3].capitalize()}-{year_val[-2:]}"
        else:
            day_val, date_str = 0, "Unknown Date"

        try:
            xls = pd.ExcelFile(file_path)
            bursa_sheet = [sheet for sheet in xls.sheet_names if "BURSA" in sheet.upper()]
            
            if bursa_sheet:
                df = pd.read_excel(file_path, sheet_name=bursa_sheet[0], header=None)
                curr_header_idx = 0
                for i in range(min(15, len(df))):
                    if df.iloc[i].astype(str).str.contains('Trade No', case=False).any():
                        curr_header_idx = i
                        break
                        
                if header_df is None:
                    header_df = df.iloc[:curr_header_idx+1].copy()
                    header_row = df.iloc[curr_header_idx]
                    for idx, val in header_row.items():
                        v_str = str(val).lower().strip()
                        if v_str == 'lots': lots_col_idx = idx
                        elif v_str in ['b/s', 'b / s', 'type', 'side', 'action', 'beli/jual', 'jual/beli']: bs_col_idx = idx
                            
                data = df.iloc[curr_header_idx+1:].copy()
                data = data[data[0].notna()]
                data = data[~data[0].astype(str).str.upper().str.contains('TOTAL')]
                
                temp_data = data.copy()
                temp_data[lots_col_idx] = pd.to_numeric(temp_data[lots_col_idx], errors='coerce').fillna(0)
                
                total_trades = len(temp_data)
                total_lots_day = temp_data[lots_col_idx].sum()
                buy_lots, sell_lots = 0, 0
                
                if bs_col_idx is not None:
                    temp_data['bs_clean'] = temp_data[bs_col_idx].astype(str).str.upper().str.strip()
                    buy_lots = temp_data.loc[temp_data['bs_clean'].str.startswith('B'), lots_col_idx].sum()
                    sell_lots = temp_data.loc[temp_data['bs_clean'].str.startswith('S'), lots_col_idx].sum()
                
                summary_rows.append({
                    'DayInt': day_val, 'Date': date_str, 'TRADE': total_trades,
                    'SELL': round(sell_lots, 3), 'BUY': round(buy_lots, 3), 'TOTAL': round(total_lots_day, 3)
                })
                data_dfs.append(data)
        except Exception:
            pass

    if not data_dfs: return None

    try:
        merged_data = pd.concat(data_dfs, ignore_index=True)
        merged_data[lots_col_idx] = pd.to_numeric(merged_data[lots_col_idx], errors='coerce')
        total_lots = round(merged_data[lots_col_idx].sum(), 3) 
        
        row_data = [np.nan] * len(merged_data.columns)
        row_data[0], row_data[lots_col_idx] = 'TOTAL LOT', total_lots
        total_row = pd.DataFrame([row_data])
        
        final_df = pd.concat([header_df, merged_data, total_row], ignore_index=True)
        
        if days and prefixes:
            min_day, max_day = min(days), max(days)
            best_prefix = max(prefixes, key=len)
            month = list(months)[0] if months else "BULAN"
            year = list(years)[0] if years else "TAHUN"
            month_mapping = {
                'JANUARY': 'JANUARI', 'FEBRUARY': 'FEBRUARI', 'MARCH': 'MARET',
                'MAY': 'MEI', 'JULY': 'JULI', 'AUGUST': 'AGUSTUS', 
                'OCTOBER': 'OKTOBER', 'DECEMBER': 'DESEMBER'
            }
            month_norm = month_mapping.get(month, month)
            merged_filename = f"{best_prefix} {min_day:02d} {month_norm} {year} (GABUNGAN).xlsx" if min_day == max_day else f"{best_prefix} {min_day:02d} - {max_day:02d} {month_norm} {year} (GABUNGAN).xlsx"
        else:
            merged_filename = "REKAP TRANSAKSI GABUNGAN TOTAL.xlsx"
            
        merged_filepath = os.path.join(download_dir, merged_filename)

        with pd.ExcelWriter(merged_filepath, engine='openpyxl') as writer:
            final_df.to_excel(writer, sheet_name='BURSA', header=False, index=False)
            if summary_rows:
                ws = writer.book.create_sheet(title="REKAP PER HARI")
                
                blue_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
                light_blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                orange_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                white_font = Font(color="FFFFFF", bold=True)
                black_bold = Font(color="000000", bold=True)
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                thick_green_border = Border(left=Side(style='medium', color="00B050"), right=Side(style='medium', color="00B050"), top=Side(style='medium', color="00B050"), bottom=Side(style='medium', color="00B050"))
                center_align = Alignment(horizontal="center", vertical="center")
                
                current_row = 2
                for row_data in sorted(summary_rows, key=lambda x: x['DayInt']):
                    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row+1, end_column=1)
                    c_trade = ws.cell(row=current_row, column=1, value="TRADE")
                    c_trade.fill, c_trade.font, c_trade.alignment, c_trade.border = blue_fill, white_font, center_align, thin_border
                    ws.cell(row=current_row+1, column=1).border = thin_border
                    
                    c_td = ws.cell(row=current_row, column=2, value="Trade date :")
                    c_td.fill, c_td.font, c_td.alignment, c_td.border = blue_fill, white_font, center_align, thin_border
                    
                    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=4)
                    c_date = ws.cell(row=current_row, column=3, value=row_data['Date'])
                    c_date.fill, c_date.font, c_date.alignment, c_date.border = blue_fill, white_font, center_align, thin_border
                    ws.cell(row=current_row, column=4).border = thin_border
                    
                    c_sell = ws.cell(row=current_row+1, column=2, value="SELL")
                    c_sell.fill, c_sell.font, c_sell.alignment, c_sell.border = light_blue_fill, black_bold, center_align, thin_border
                    c_buy = ws.cell(row=current_row+1, column=3, value="BUY")
                    c_buy.fill, c_buy.font, c_buy.alignment, c_buy.border = orange_fill, black_bold, center_align, thin_border
                    c_tot = ws.cell(row=current_row+1, column=4, value="BUY + SELL")
                    c_tot.fill, c_tot.font, c_tot.alignment, c_tot.border = green_fill, black_bold, center_align, thick_green_border
                    
                    c_tv = ws.cell(row=current_row+2, column=1, value=row_data['TRADE'])
                    c_tv.fill, c_tv.font, c_tv.alignment, c_tv.border = light_blue_fill, black_bold, center_align, thin_border
                    c_sv = ws.cell(row=current_row+2, column=2, value=row_data['SELL'])
                    c_sv.fill, c_sv.alignment, c_sv.border, c_sv.number_format = light_blue_fill, center_align, thin_border, '#,##0.00'
                    c_bv = ws.cell(row=current_row+2, column=3, value=row_data['BUY'])
                    c_bv.fill, c_bv.alignment, c_bv.border, c_bv.number_format = light_blue_fill, center_align, thin_border, '#,##0.00'
                    c_totv = ws.cell(row=current_row+2, column=4, value=row_data['TOTAL'])
                    c_totv.fill, c_totv.font, c_totv.alignment, c_totv.border, c_totv.number_format = light_blue_fill, black_bold, center_align, thick_green_border, '#,##0.00'
                    current_row += 4
                    
                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 20
                ws.column_dimensions['C'].width = 20
                ws.column_dimensions['D'].width = 20

        return merged_filename
    except Exception:
        return None

def buat_template_ldk(nama_bulan="Bulan", tahun=2026, daftar_cabang=[]):
    wb = Workbook()
    font_bold = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(top=Side(style="thin"), left=Side(style="thin"), right=Side(style="thin"), bottom=Side(style="thin"))
    
    def apply_style(ws, row, col, bold=False, center=True, border=True):
        cell = ws.cell(row=row, column=col)
        if bold: cell.font = font_bold
        if center: cell.alignment = align_center
        if border: cell.border = thin_border
        return cell

    # SHEET VOL
    ws = wb.active
    ws.title = "Vol"
    ws['A1'], ws['A2'], ws['A3'], ws['A4'] = "Form 7", "Rekapitulasi Volume Transaksi (Lot)", "PT. CENTRAL CAPITAL  FUTURES", tahun 
    
    headers_vol = {'A6': "No.", 'B6': "Lokasi", 'C6': nama_bulan, 'C7': "Kontrak Berjangka", 'E7': "Kontrak Derivatif Lainnya", 'G7': "Jumlah", 'C8': "Pertambangan", 'D8': "Pertanian Perkebunan", 'E8': "SPA", 'F8': "PALN"}
    for pos, val in headers_vol.items(): ws[pos] = val
    
    ws.merge_cells('A6:A8'); ws.merge_cells('B6:B8'); ws.merge_cells('C6:G6'); ws.merge_cells('C7:D7'); ws.merge_cells('E7:F7'); ws.merge_cells('G7:G8')
    for r in range(6, 9):
        for c in range(1, 8): apply_style(ws, r, c, bold=True)
        
    cabang_list = ["PUSAT"] + daftar_cabang
    data_vol = []
    for i, cab in enumerate(cabang_list, start=1):
        r_idx = 8 + i
        data_vol.append([i, cab, 0, 0, 0, 0, f"=SUM(D{r_idx}:F{r_idx})"])

    for r_idx, row_data in enumerate(data_vol, start=9):
        for c_idx, val in enumerate(row_data, start=1):
            apply_style(ws, r_idx, c_idx, center=(c_idx == 1 or c_idx >= 3)).value = val

    total_row = 9 + len(data_vol) + 1
    ws[f'A{total_row}'] = "TOTAL"
    ws.merge_cells(f'A{total_row}:B{total_row}')
    apply_style(ws, total_row, 1, bold=True)
    apply_style(ws, total_row, 2, bold=True)
    for col, letter in enumerate(['C', 'D', 'E', 'F', 'G'], start=3):
        apply_style(ws, total_row, col, bold=True).value = f"=SUM({letter}9:{letter}{total_row-2})"
        
    ws[f'A{total_row + 2}'] = "Catatan"
    ws[f'A{total_row + 3}'] = "*)Isi lokasi dengan nama kantor cabang, pastikan nama yang terdapat pula di tab domisili"
    ws[f'A{total_row + 4}'] = "*)pastikan dalam satu baris semua colom terisi"
    
    ws.column_dimensions['B'].width, ws.column_dimensions['C'].width, ws.column_dimensions['D'].width = 20, 15, 20
    ws.column_dimensions['E'].width, ws.column_dimensions['F'].width, ws.column_dimensions['G'].width = 15, 15, 15

    # SHEET KOM
    wk = wb.create_sheet("Kom")
    wk['A1'], wk['A2'], wk['A3'], wk['A4'] = "Form 8", "Rekapitulasi Komisi Transaksi", "PT. CENTRAL CAPITAL  FUTURES", tahun 
    
    headers_kom = {'A6': "No.", 'B6': "Jenis Kontrak", 'D6': nama_bulan, 'D7': "Vol (Lot)", 'E7': "Komisi (per Lot)", 'F7': "Total Komisi"}
    for pos, val in headers_kom.items(): wk[pos] = val
    
    wk.merge_cells('A6:A7'); wk.merge_cells('B6:C7'); wk.merge_cells('D6:F6')
    for r in range(6, 8):
        for c in range(1, 7): apply_style(wk, r, c, bold=True)
        
    wk['A8'], wk['B8'] = 1, "Kontrak Berjangka"
    wk['D8'], wk['E8'], wk['F8'] = "=SUM(D9:D9)", "=SUM(E9:E9)", "=SUM(F9:F9)"
    wk.merge_cells('B8:C8')
    for c in range(1, 7): apply_style(wk, 8, c, bold=True)
    
    for c_idx, val in enumerate(["-", "KGE", 0, 0, "=D9*E9"], start=2): apply_style(wk, 9, c_idx).value = val
    apply_style(wk, 9, 1)

    wk['A10'], wk['B10'] = 2, "Kontrak Derivatif Lainnya"
    wk['D10'], wk['E10'], wk['F10'] = "=D11+D75", "=E11+E75", "=F11+F75"
    wk.merge_cells('B10:C10')
    for c in range(1, 7): apply_style(wk, 10, c, bold=True)
    
    wk['B11'], wk['C11'] = "-", "SPA"
    wk['D11'], wk['E11'], wk['F11'] = "=SUM(D12:D74)", "=SUM(E12:E74)", "=SUM(F12:F74)"
    for c in range(1, 7): apply_style(wk, 11, c, bold=True)
    
    commodities = [
        "JPK50_BBJ", "HKJ50_BBJ", "KRK50_BBJ", "EU1010_BBJ", "UJ1010_BBJ", "GU1010_BBJ", "UC1010_BBJ", "AU1010_BBJ", 
        "UK1010_BBJ", "NU1010_BBJ", "EJ1H10_BBJ", "GJ1H10_BBJ", "GA1H10", "CJ1H10_BBJ", "AJ1H10_BBJ", "EG1H10_BBJ", 
        "EN1010_BBJ", "EC1010_BBJ", "GC1010_BBJ", "AN1010_BBJ", "XAG10_BBJ", "XUL10_BBJ", "CLSK10_BBJ", "EU1212_BBJ", 
        "UJ1212_BBJ", "GU1212_BBJ", "UC1212_BBJ", "AU1212_BBJ", "UK1212_BBJ", "NU1212_BBJ", "EJ1H12_BBJ", "GJ1H12_BBJ", 
        "GA1H12", "CJ1H12_BBJ", "AJ1H12_BBJ", "EG1H12_BBJ", "EN1212_BBJ", "EC1212_BBJ", "GC1212_BBJ", "AN1212_BBJ", 
        "XAG12_BBJ", "XUL12_BBJ", "CLSK12_BBJ", "EU1414_BBJ", "UJ1414_BBJ", "GU1414_BBJ", "UC1414_BBJ", "AU1414_BBJ", 
        "UK1414_BBJ", "NU1414_BBJ", "EJ1H14_BBJ", "GJ1H14_BBJ", "GA1H14", "CJ1H14_BBJ", "AJ1H14_BBJ", "EG1H14_BBJ", 
        "EN1414_BBJ", "EC1414_BBJ", "GC1414_BBJ", "AN1414_BBJ", "XAG14_BBJ", "XUL14_BBJ", "CLSK14_BBJ"
    ]
    
    r_idx = 12
    for comm in commodities:
        apply_style(wk, r_idx, 1)
        apply_style(wk, r_idx, 2).value = "-"
        apply_style(wk, r_idx, 3).value = comm
        apply_style(wk, r_idx, 4).value = 0
        apply_style(wk, r_idx, 5).value = 20
        apply_style(wk, r_idx, 6).value = f"=D{r_idx}*E{r_idx}"
        r_idx += 1
        
    for c_idx, val in enumerate(["", "-", "PALN", "=(0)*2", f"=E{r_idx+1}", f"=F{r_idx+1}"], start=1): apply_style(wk, r_idx, c_idx).value = val
    for c_idx, val in enumerate(["", "-", "", "=(0)*2", 0, f"=D{r_idx+1}*E{r_idx+1}"], start=1): apply_style(wk, r_idx+1, c_idx).value = val
    
    apply_style(wk, r_idx+2, 1); apply_style(wk, r_idx+2, 2)
    apply_style(wk, r_idx+2, 3, bold=True).value = "TOTAL"
    apply_style(wk, r_idx+2, 4, bold=True).value = "=D8+D10"
    apply_style(wk, r_idx+2, 5, bold=True).value = "=E8+E10"
    apply_style(wk, r_idx+2, 6, bold=True).value = "=F8+F10"
    
    wk.cell(row=r_idx+4, column=2).value = "Catatan :"
    wk.cell(row=r_idx+5, column=2).value = "Kode kontrak harus diisi jika tidak diisi maka data tidakakan dibaca"
    wk.cell(row=r_idx+6, column=2).value = "Data harus berurut jika terselip baris yang kosong maka baris berikunya tidak dibaca"
    
    wk.column_dimensions['B'].width, wk.column_dimensions['C'].width = 5, 25
    wk.column_dimensions['D'].width, wk.column_dimensions['E'].width, wk.column_dimensions['F'].width = 15, 18, 18
    
    return wb

PREFIX_KLIEN       = "CC"
KOMISI_DEFAULT     = 20
UTAMAKAN_REVISI    = True
COL_COMMODITY, COL_LOT, COL_K, COL_N = 3, 6, 10, 13

def normalisasi_kode(v):
    s = str(v).replace('\u00a0', ' ').strip()
    if re.fullmatch(r'\d+\.0+', s): s = s.split('.')[0]
    return s

def ekstrak_akun(row):
    kandidat = [str(row[COL_K]).strip(), str(row[COL_N]).strip()]
    for v in kandidat:
        if v.upper().startswith(PREFIX_KLIEN) and (mm := re.search(r'(\d{5,})', v)): return mm.group(1)
    for v in kandidat:
        if mm := re.search(r'(\d{5,})', v): return mm.group(1)
    return None

def cari_baris(ws, teks, kolom=3):
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, kolom).value).strip().upper() == teks.upper(): return r
    return None

def sisipkan_komoditas(ws, kode, lot):
    r_spa = cari_baris(ws, 'SPA')
    mm = re.match(r'=SUM\(D(\d+):D(\d+)\)', str(ws.cell(r_spa, 4).value))
    if not mm: return None
    awal, akhir = int(mm.group(1)), int(mm.group(2))
    baru = akhir + 1
    ws.insert_rows(baru)
    for c in range(1, 7): ws.cell(baru, c)._style = copy.copy(ws.cell(baru - 1, c)._style)
    ws.cell(baru, 2).value = '-'
    ws.cell(baru, 3).value = kode
    ws.cell(baru, 4).value = float(lot)
    ws.cell(baru, 5).value = KOMISI_DEFAULT
    ws.cell(baru, 6).value = f'=D{baru}*E{baru}'
    for c, h in ((4, 'D'), (5, 'E'), (6, 'F')): ws.cell(r_spa, c).value = f'=SUM({h}{awal}:{h}{baru})'
    r_paln, r_lain, r_total  = cari_baris(ws, 'PALN'), cari_baris(ws, 'Kontrak Derivatif Lainnya', kolom=2), cari_baris(ws, 'TOTAL')
    
    if r_paln:
        ws.cell(r_paln, 5).value = f'=E{r_paln + 1}'
        ws.cell(r_paln, 6).value = f'=F{r_paln + 1}'
        ws.cell(r_paln + 1, 6).value = f'=D{r_paln + 1}*E{r_paln + 1}'
        if r_lain:
            for c, h in ((4, 'D'), (5, 'E'), (6, 'F')): ws.cell(r_lain, c).value = f'={h}{r_spa}+{h}{r_paln}'
    if r_total and r_lain:
        for c, h in ((4, 'D'), (5, 'E'), (6, 'F')): ws.cell(r_total, c).value = f'={h}8+{h}{r_lain}'
    return baru

# ============================================================
# 3. GUI STREAMLIT MAIN
# ============================================================

st.title("✨ Dashboard Automasi LDK")
st.markdown("Pusat kendali untuk mengunduh rekap dari email dan membangun Form LDK secara otomatis.")

# --- PENGATURAN GLOBAL (EXPANDER) ---
with st.expander("⚙️ Konfigurasi Logika Address Cabang (Klik untuk membuka)", expanded=False):
    st.write("Atur logika pemetaan *address* cabang di bawah ini. Pusat akan ditangkap otomatis dari sisa alamat yang tidak terdaftar.")
    
    default_cabang = pd.DataFrame(load_config())
    df_cabang_edited = st.data_editor(
        default_cabang, 
        num_rows="dynamic", 
        use_container_width=True,
        hide_index=True 
    )

    clean_data = []
    for _, row in df_cabang_edited.iterrows():
        nama = str(row.get('Nama Cabang', '')).strip()
        logika = str(row.get('Logika Address', '')).strip()
        if nama and logika and nama.lower() != 'nan' and logika.lower() != 'nan':
            clean_data.append({"Nama Cabang": nama, "Logika Address": logika})
    save_config(clean_data)

# --- TABS UTAMA ---
tab1, tab2 = st.tabs(["📥 1. Downloader Lampiran Email", "📊 2. Ekstraksi & Builder LDK"])

# ------------------------------------------------------------
# TAB 1: DOWNLOADER EMAIL
# ------------------------------------------------------------
with tab1:
    st.subheader("Unduh File Gabungan Bursa dari Yahoo")
    with st.container(border=True):
        col1, col2 = st.columns(2)
        with col1:
            email_acc = st.text_input("Akun Email Yahoo", value="dealingccf_bbj@yahoo.com")
            app_pass = "krrrgdmbdoxorfbv" # Password aplikasi
        with col2:
            imap_server = st.text_input("Server IMAP", value="imap.mail.yahoo.com")
            subject_input = st.text_input("Subjek Pencarian (pisahkan koma jika banyak)", value="REKAP TRANSAKSI Shift III JANUARI 2026")

        if st.button("🚀 Mulai Ekstraksi Email", type="primary", use_container_width=True):
            if not subject_input:
                st.warning("Pencarian dibatalkan karena subjek kosong.")
            else:
                DOWNLOAD_DIR = "lampiran_email_temp"
                if os.path.exists(DOWNLOAD_DIR): shutil.rmtree(DOWNLOAD_DIR)
                os.makedirs(DOWNLOAD_DIR, exist_ok=True)

                queries = [q.strip() for q in subject_input.split(",") if q.strip()]
                dynamic_zip_filename = f"{clean_filename(queries[0])[:50] or 'Hasil_Lampiran'}.zip"

                with st.status("Sedang memproses email...", expanded=True) as status:
                    try:
                        st.write("Menghubungkan ke server IMAP...")
                        mail = imaplib.IMAP4_SSL(imap_server)
                        mail.login(email_acc, app_pass)
                        mail.select('"INBOX"')
                        
                        shift3_ids, shift2_ids = set(), set()
                        st.write("Mencari kecocokan subjek...")
                        for query in queries:
                            shift3_ids.update(do_imap_search(mail, query))
                            if "Shift III" in query:
                                shift2_ids.update(do_imap_search(mail, query.replace("Shift III", "Shift II")))

                        shift2_ids = shift2_ids - shift3_ids
                        total_emails = len(shift3_ids) + len(shift2_ids)

                        if total_emails == 0:
                            status.update(label="Tidak ada email ditemukan.", state="error")
                        else:
                            download_count = 0
                            downloaded_dates = set()
                            progress_bar = st.progress(0)

                            def process_emails(email_ids, is_fallback=False):
                                global download_count
                                for idx, eid in enumerate(list(email_ids)):
                                    res, msg_data = mail.fetch(eid, "(RFC822)")
                                    if res == "OK":
                                        for response_part in msg_data:
                                            if isinstance(response_part, tuple):
                                                msg = email.message_from_bytes(response_part[1])
                                                safe_subject = clean_filename(decode_mime_header(msg.get("Subject", "Tanpa_Subjek")))

                                                if msg.is_multipart():
                                                    for part in msg.walk():
                                                        if part.get_content_maintype() == 'multipart' or part.get('Content-Disposition') is None: continue
                                                        original_filename = part.get_filename()
                                                        if original_filename:
                                                            _, ext = os.path.splitext(decode_mime_header(original_filename))
                                                            filename = clean_filename(f"{safe_subject}{ext}")
                                                            date_key = extract_date_key(filename)
                                                            
                                                            if date_key:
                                                                if is_fallback and date_key in downloaded_dates: continue
                                                                if not is_fallback: downloaded_dates.add(date_key)

                                                            filepath = os.path.join(DOWNLOAD_DIR, filename)
                                                            base, counter = os.path.splitext(filename)[0], 1
                                                            while os.path.exists(filepath):
                                                                filepath = os.path.join(DOWNLOAD_DIR, f"{base}_{counter}{ext}")
                                                                counter += 1

                                                            with open(filepath, "wb") as f: f.write(part.get_payload(decode=True))
                                                            download_count += 1
                                    progress_bar.progress((idx + 1) / total_emails)

                            process_emails(shift3_ids, False)
                            process_emails(shift2_ids, True)
                            
                            st.write("Filter file 'Revisi'...")
                            all_downloaded = glob.glob(os.path.join(DOWNLOAD_DIR, "*.*"))
                            files_by_date = {}
                            for fpath in all_downloaded:
                                d_key = extract_date_key(os.path.basename(fpath))
                                if d_key: files_by_date.setdefault(f"{d_key}_{os.path.splitext(fpath)[1].lower()}", []).append(fpath)

                            for flist in files_by_date.values():
                                if len(flist) > 1:
                                    revisi_files = sorted([f for f in flist if 'revisi' in os.path.basename(f).lower()])
                                    if revisi_files:
                                        kept_file = revisi_files[-1]
                                        for f in flist:
                                            if f != kept_file:
                                                try: os.remove(f); download_count -= 1
                                                except: pass

                            if download_count > 0:
                                merge_downloaded_excel(DOWNLOAD_DIR)
                                shutil.make_archive(dynamic_zip_filename.replace('.zip', ''), 'zip', DOWNLOAD_DIR)
                                status.update(label=f"Selesai! {download_count} file berhasil diunduh.", state="complete")
                                with open(dynamic_zip_filename, "rb") as f:
                                    st.download_button("📥 Unduh Hasil Ekstraksi (ZIP)", data=f, file_name=dynamic_zip_filename, mime="application/zip", type="primary")
                            else:
                                status.update(label="Email ditemukan, namun tidak ada lampiran valid.", state="error")
                    except Exception as e:
                        status.update(label=f"Error: {e}", state="error")
                    finally:
                        try: mail.close(); mail.logout()
                        except: pass

# ------------------------------------------------------------
# TAB 2: EKSTRAK & BUILDER LDK
# ------------------------------------------------------------
with tab2:
    st.subheader("Otomatisasi Peta Akun & Format LDK")
    
    with st.container(border=True):
        col_a, col_b = st.columns(2)
        with col_a:
            zip_file_upload = st.file_uploader("📦 1. Upload ZIP Rekap Transaksi", type=["zip"])
        with col_b:
            acc_file_upload = st.file_uploader("📋 2. Upload File List ACC (.xlsx)", type=["xlsx"])

        if st.button("🔨 Proses & Buat File LDK", type="primary", use_container_width=True):
            if not (zip_file_upload and acc_file_upload):
                st.error("Mohon lengkapi unggahan file ZIP dan List ACC terlebih dahulu.")
            else:
                EXTRACT_DIR = "rekap_extracted_temp"
                if os.path.exists(EXTRACT_DIR): shutil.rmtree(EXTRACT_DIR)
                os.makedirs(EXTRACT_DIR, exist_ok=True)
                
                with open("temp_rekap.zip", "wb") as f: f.write(zip_file_upload.getbuffer())
                with open("temp_acc.xlsx", "wb") as f: f.write(acc_file_upload.getbuffer())

                with st.status("Memproses injeksi data...", expanded=True) as status:
                    try:
                        with zipfile.ZipFile("temp_rekap.zip", 'r') as z: z.extractall(EXTRACT_DIR)

                        file_rekaps = sorted([os.path.join(r, f) for r, d, fl in os.walk(EXTRACT_DIR) if '__MACOSX' not in r for f in fl if f.endswith(('.xlsx', '.xls')) and not f.startswith(('~', '.')) and 'GABUNGAN' not in f.upper()])

                        if UTAMAKAN_REVISI:
                            POLA_REVISI = r'\b(REV|REVISI|REVISED|PERBAIKAN|FIX|UPDATE)\b'
                            is_revisi = lambda p: re.search(POLA_REVISI, os.path.basename(p).upper()) is not None
                            kunci = lambda p: re.sub(r'\s+', ' ', re.sub(POLA_REVISI, ' ', re.sub(r'\.(XLSX|XLS)$', '', os.path.basename(p).upper()))).strip()
                            ada_revisi = {kunci(p) for p in file_rekaps if is_revisi(p)}
                            file_rekaps = [p for p in file_rekaps if not (not is_revisi(p) and kunci(p) in ada_revisi)]

                        if not file_rekaps:
                            status.update(label="Data rekap kosong / tidak valid.", state="error")
                        else:
                            nama_bulan, tahun = "Bulan", None
                            for bln in ['JANUARI', 'FEBRUARI', 'MARET', 'APRIL', 'MEI', 'JUNI', 'JULI', 'AGUSTUS', 'SEPTEMBER', 'OKTOBER', 'NOVEMBER', 'DESEMBER']:
                                if bln in os.path.basename(file_rekaps[0]).upper():
                                    nama_bulan = bln.capitalize()
                                    break
                            m = re.search(r'(20\d{2})', os.path.basename(file_rekaps[0]))
                            if m: tahun = int(m.group(1))

                            st.write(f"📁 Membaca {len(file_rekaps)} file untuk periode **{nama_bulan} {tahun or ''}**...")

                            dict_cabang, daftar_cabang_tambahan = {}, []
                            for item in clean_data:
                                nama_cabang = item['Nama Cabang'].upper()
                                dict_cabang[item['Logika Address']] = nama_cabang
                                if nama_cabang not in daftar_cabang_tambahan: daftar_cabang_tambahan.append(nama_cabang)

                            data_dfs = []
                            for f in file_rekaps:
                                try:
                                    xls = pd.ExcelFile(f)
                                    bursa_sheet = [s for s in xls.sheet_names if "BURSA" in s.upper()]
                                    data_dfs.append(pd.read_excel(f, sheet_name=bursa_sheet[0] if bursa_sheet else 0, header=None))
                                except Exception: pass
                            
                            if not data_dfs:
                                status.update(label="Gagal mengekstrak isi Excel.", state="error")
                            else:
                                df = pd.concat(data_dfs, ignore_index=True)
                                df = df.sort_values(by=0, key=lambda x: x.astype(str)).reset_index(drop=True)
                                df['Commodity'] = df[COL_COMMODITY].astype(str).str.replace('\u00a0', ' ').str.strip()
                                df = df[df['Commodity'] != 'Commodity']
                                df['Lot'] = pd.to_numeric(df[COL_LOT], errors='coerce').fillna(0)
                                df['ACC'] = df.apply(ekstrak_akun, axis=1)
                                
                                df_valid = df.dropna(subset=['ACC']).copy()
                                acc = pd.read_excel("temp_acc.xlsx", header=None)
                                acc['Login']   = acc[0].map(normalisasi_kode)
                                acc['Address'] = acc[5].map(lambda v: str(v).replace('\u00a0', ' ').strip())
                                dict_address = dict(zip(acc['Login'], acc['Address']))

                                df_valid['ACC'] = df_valid['ACC'].map(normalisasi_kode)
                                df_valid['Address'] = df_valid['ACC'].map(dict_address)

                                if df_valid['Address'].notna().sum() == 0:
                                    status.update(label="FATAL: Akun tidak cocok dengan List ACC.", state="error")
                                else:
                                    df_valid['Cabang'] = df_valid['Address'].map(dict_cabang).fillna('PUSAT')
                                    lot_per_cabang = df_valid.groupby('Cabang')['Lot'].sum().to_dict()
                                    
                                    df_kom = df_valid.groupby('Commodity')['Lot'].sum().reset_index()
                                    df_kom['Lot'] = df_kom['Lot'].round(4)

                                    file_output = f"Output_LDK_{nama_bulan}{tahun or ''}.xlsx"
                                    
                                    wb = buat_template_ldk(nama_bulan=nama_bulan, tahun=(tahun if tahun else 2026), daftar_cabang=daftar_cabang_tambahan)
                                    ws = wb['Vol']
                                    semua_cabang_urut = ['PUSAT'] + daftar_cabang_tambahan
                                    
                                    for row in ws.iter_rows(min_row=9, max_row=8 + len(semua_cabang_urut)):
                                        if row[1].value:
                                            lok = str(row[1].value).upper().replace(" ", "")
                                            for cab in semua_cabang_urut:
                                                if cab.replace(" ", "") == lok: row[4].value = round(lot_per_cabang.get(cab, 0.0), 4)

                                    wk = wb['Kom']
                                    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Kuning pastel yang nyaman di mata

                                    for _, d in df_kom.iterrows():
                                        kode, lot = str(d['Commodity']).strip(), float(d['Lot'])
                                        if kode in ("nan", ""): continue
                                        
                                        peta = {str(wk.cell(r, 3).value).strip().upper(): r for r in range(2, wk.max_row + 1) if str(wk.cell(r, 3).value).strip() and str(wk.cell(r, 3).value).strip().lower() != 'none'}
                                        baris = peta.get(kode.upper()) or next((r for k_ldk, r in peta.items() if kode.upper() in k_ldk), None) or sisipkan_komoditas(wk, kode, lot)
                                        
                                        if baris:
                                            wk.cell(baris, 4).value = lot
                                            if lot > 0:
                                                for col in range(1, 7): wk.cell(row=baris, column=col).fill = yellow_fill

                                    wb.save(file_output)

                                    # UX: Tampilkan hasil metrik di atas tombol Download
                                    st.write("---")
                                    st.subheader("📈 Ringkasan Volume (Lot)")
                                    
                                    metric_cols = st.columns(len(semua_cabang_urut) + 1)
                                    for i, cab in enumerate(semua_cabang_urut):
                                        metric_cols[i].metric(label=f"🏙️ {cab}", value=f"{round(lot_per_cabang.get(cab, 0.0), 4):,}")
                                    metric_cols[-1].metric(label="📊 TOTAL KESELURUHAN", value=f"{round(sum(lot_per_cabang.values()), 4):,}")

                                    # -----------------------------------------------------------------
                                    # BARU: FITUR PRATINJAU (PREVIEW) TABEL HASIL
                                    # -----------------------------------------------------------------
                                    with st.expander("👀 Klik di sini untuk Pratinjau Rincian Tabel LDK", expanded=True):
                                        col_prev1, col_prev2 = st.columns(2)
                                        
                                        with col_prev1:
                                            st.markdown("**1️⃣ Rekapitulasi Volume Transaksi per Lokasi**")
                                            # Membangun dataframe rekap per cabang
                                            df_prev_vol = pd.DataFrame(list(lot_per_cabang.items()), columns=["Lokasi Cabang", "Total Lot"])
                                            df_prev_vol.loc[len(df_prev_vol)] = ["TOTAL KESELURUHAN", sum(lot_per_cabang.values())]
                                            st.dataframe(df_prev_vol, hide_index=True, use_container_width=True)
                                            
                                        with col_prev2:
                                            st.markdown("**2️⃣ Rincian Lot per Komoditas Aktif**")
                                            # Memfilter hanya komoditas yang nilai lot-nya > 0
                                            df_prev_kom = df_kom[df_kom['Lot'] > 0].copy()
                                            if df_prev_kom.empty:
                                                st.info("Tidak ada transaksi / lot bernilai 0.")
                                            else:
                                                df_prev_kom.columns = ["Jenis Komoditas", "Total Lot"]
                                                df_prev_kom.loc[len(df_prev_kom)] = ["TOTAL", df_prev_kom['Total Lot'].sum()]
                                                st.dataframe(df_prev_kom, hide_index=True, use_container_width=True)
                                    st.write("---")

                                    status.update(label="Selesai! LDK berhasil dibuat.", state="complete")
                                    
                                    with open(file_output, "rb") as f:
                                        st.download_button("🎉 Unduh Template LDK Final (Excel)", data=f, file_name=file_output, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")

                    except Exception as e:
                        status.update(label=f"Terjadi kesalahan saat memproses: {e}", state="error")
